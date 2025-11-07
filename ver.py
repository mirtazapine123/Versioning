import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import sqlite3
from datetime import datetime
import os
import base64
from PIL import Image, ImageGrab, ImageTk
import io
from difflib import SequenceMatcher
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class MachineTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema Tracciamento Modifiche Macchine")
        self.root.geometry("1400x800")
        self.current_attachments = []
        self.preview_widgets = []
        self.init_database()
        self.create_widgets()        
        self.load_all_records()
    
    def init_database(self):
        self.conn = sqlite3.connect('macchine_tracker.db')
        self.cursor = self.conn.cursor()
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS interventi (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_ora TEXT NOT NULL,
                macchina TEXT NOT NULL,
                operatore TEXT NOT NULL,
                categoria TEXT NOT NULL,
                problema TEXT NOT NULL,
                soluzione TEXT NOT NULL
            )
        ''')
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS allegati (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                intervento_id INTEGER NOT NULL,
                nome_file TEXT NOT NULL,
                tipo_file TEXT NOT NULL,
                contenuto BLOB NOT NULL,
                FOREIGN KEY (intervento_id) REFERENCES interventi(id) ON DELETE CASCADE
            )
        ''')
        
        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='immagini'")
        if self.cursor.fetchone():
            try:
                self.cursor.execute('''
                    INSERT INTO allegati (intervento_id, nome_file, tipo_file, contenuto)
                    SELECT intervento_id, nome_file, 'image', immagine FROM immagini
                ''')
                self.cursor.execute('DROP TABLE immagini')
                self.conn.commit()
            except:
                pass
        
        self.conn.commit()
    
    def create_widgets(self):
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.tab_insert = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_insert, text='Nuovo Intervento')
        self.create_insert_tab()
        
        self.tab_search = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_search, text='Ricerca e Storico')
        self.create_search_tab()
        
        self.tab_ai = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_ai, text='Assistente IA')
        self.create_ai_tab()
        
        self.tab_stats = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stats, text='Statistiche')
        self.create_stats_tab()
    
    def create_insert_tab(self):
        main_frame = ttk.Frame(self.tab_insert, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.tab_insert.columnconfigure(0, weight=1)
        self.tab_insert.rowconfigure(0, weight=1)

        left_frame = ttk.LabelFrame(main_frame, text="Dati Intervento", padding="10")
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        main_frame.columnconfigure(0, weight=2)
        main_frame.rowconfigure(0, weight=1)
        
        ttk.Label(left_frame, text="Macchina:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.macchina_entry = ttk.Entry(left_frame, width=50)
        self.macchina_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(left_frame, text="Operatore:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.operatore_entry = ttk.Entry(left_frame, width=50)
        self.operatore_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(left_frame, text="Categoria:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.categoria_combo = ttk.Combobox(left_frame, width=47, state="readonly")
        self.categoria_combo['values'] = ('Software', 'Elettrico', 'Pneumatico', 'Meccanico', 'Manutenzione', 'Altro')
        self.categoria_combo.current(0)
        self.categoria_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(left_frame, text="Problema:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.problema_text = scrolledtext.ScrolledText(left_frame, width=50, height=8, wrap=tk.WORD)
        self.problema_text.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(left_frame, text="Soluzione:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.soluzione_text = scrolledtext.ScrolledText(left_frame, width=50, height=8, wrap=tk.WORD)
        self.soluzione_text.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5)
        
        button_frame = ttk.Frame(left_frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="💾 Salva Intervento", command=self.save_record).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="🗑️ Pulisci Campi", command=self.clear_fields).pack(side=tk.LEFT, padx=5)
        
        left_frame.columnconfigure(1, weight=1)
        
        right_frame = ttk.LabelFrame(main_frame, text="Allegati (Immagini e Documenti)", padding="10")
        right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))
        main_frame.columnconfigure(1, weight=1)
        
        attach_buttons = ttk.Frame(right_frame)
        attach_buttons.pack(fill='x', pady=(0, 10))
        
        ttk.Button(attach_buttons, text="📷 Screenshot", command=self.take_screenshot).pack(side=tk.LEFT, padx=2)
        ttk.Button(attach_buttons, text="🖼️ Immagine", command=self.load_image).pack(side=tk.LEFT, padx=2)
        ttk.Button(attach_buttons, text="📄 File TXT", command=self.load_txt_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(attach_buttons, text="📝 File DOCX", command=self.load_docx_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(attach_buttons, text="❌ Rimuovi", command=self.remove_attachment).pack(side=tk.LEFT, padx=2)
        
        preview_frame = ttk.Frame(right_frame)
        preview_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(preview_frame, height=500)
        scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=canvas.yview)
        self.preview_container = ttk.Frame(canvas)
        
        self.preview_container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.preview_container, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def create_search_tab(self):
        main_frame = ttk.Frame(self.tab_search, padding="10")
        main_frame.pack(fill='both', expand=True)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        search_frame = ttk.LabelFrame(main_frame, text="Ricerca", padding="10")
        search_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        search_frame.columnconfigure(1, weight=1)
        
        ttk.Label(search_frame, text="Cerca:").grid(row=0, column=0, padx=(0, 5))
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        self.search_entry.bind('<KeyRelease>', lambda e: self.search_records())
        
        ttk.Button(search_frame, text="🔍 Cerca", command=self.search_records).grid(row=0, column=2, padx=5)
        ttk.Button(search_frame, text="📋 Mostra Tutti", command=self.load_all_records).grid(row=0, column=3, padx=5)
        ttk.Button(search_frame, text="📊 Export Excel", command=self.export_to_excel).grid(row=0, column=4, padx=5)
        
        results_frame = ttk.Frame(main_frame)
        results_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        tree_scroll = ttk.Scrollbar(results_frame)
        tree_scroll.pack(side='right', fill='y')
        
        self.tree = ttk.Treeview(results_frame, yscrollcommand=tree_scroll.set, selectmode='browse')
        self.tree.pack(side='left', fill='both', expand=True)
        tree_scroll.config(command=self.tree.yview)
        
        self.tree['columns'] = ('Data', 'Macchina', 'Operatore', 'Categoria', 'Problema')
        self.tree.column('#0', width=0, stretch=tk.NO)
        self.tree.column('Data', anchor=tk.W, width=130)
        self.tree.column('Macchina', anchor=tk.W, width=120)
        self.tree.column('Operatore', anchor=tk.W, width=100)
        self.tree.column('Categoria', anchor=tk.W, width=100)
        self.tree.column('Problema', anchor=tk.W, width=400)
        
        self.tree.heading('Data', text='Data/Ora')
        self.tree.heading('Macchina', text='Macchina')
        self.tree.heading('Operatore', text='Operatore')
        self.tree.heading('Categoria', text='Categoria')
        self.tree.heading('Problema', text='Problema')
        
        self.tree.bind('<Double-1>', self.show_details)
        
        details_frame = ttk.LabelFrame(main_frame, text="Dettagli Intervento", padding="10")
        details_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.details_text = scrolledtext.ScrolledText(details_frame, height=10, wrap=tk.WORD, state='disabled')
        self.details_text.pack(fill='both', expand=True)
        
        btn_frame = ttk.Frame(details_frame)
        btn_frame.pack(pady=(10, 0))
        
        ttk.Button(btn_frame, text="📎 Visualizza Allegati", command=self.view_attachments).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ Elimina Intervento", command=self.delete_record).pack(side=tk.LEFT, padx=5)
    
    def create_ai_tab(self):
        main_frame = ttk.Frame(self.tab_ai, padding="10")
        main_frame.pack(fill='both', expand=True)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        header = ttk.Label(main_frame, text="🤖 Assistente IA - Trova Soluzioni Simili", 
                          font=('Arial', 14, 'bold'))
        header.grid(row=0, column=0, pady=(0, 10))
        
        input_frame = ttk.LabelFrame(main_frame, text="Descrivi il Problema", padding="10")
        input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        input_frame.columnconfigure(0, weight=1)
        
        self.ai_question = scrolledtext.ScrolledText(input_frame, height=4, wrap=tk.WORD)
        self.ai_question.pack(fill='x', pady=(0, 10))
        
        btn_frame = ttk.Frame(input_frame)
        btn_frame.pack()
        
        ttk.Button(btn_frame, text="🔍 Cerca Soluzioni Simili", command=self.ai_find_solutions).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ Pulisci", command=lambda: self.ai_question.delete('1.0', tk.END)).pack(side=tk.LEFT, padx=5)
        
        results_frame = ttk.LabelFrame(main_frame, text="Soluzioni Trovate", padding="10")
        results_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        self.ai_results = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, state='disabled')
        self.ai_results.pack(fill='both', expand=True)
    
    def create_stats_tab(self):
        main_frame = ttk.Frame(self.tab_stats, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        ttk.Button(main_frame, text="🔄 Aggiorna Statistiche", command=self.update_statistics).pack(pady=10)
        
        self.stats_container = ttk.Frame(main_frame)
        self.stats_container.pack(fill='both', expand=True)
        
        self.update_statistics()
    
    def take_screenshot(self):
        self.root.withdraw()
        self.root.after(500, self._capture_screen)
    
    def _capture_screen(self):
        try:
            screenshot = ImageGrab.grab()
            
            img_byte_arr = io.BytesIO()
            screenshot.save(img_byte_arr, format='PNG')
            img_data = img_byte_arr.getvalue()
            
            self.current_attachments.append({
                'name': f'screenshot_{len(self.current_attachments)+1}.png',
                'type': 'image',
                'data': img_data
            })
            
            self.update_attachments_preview()
            
        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante lo screenshot: {e}")
        finally:
            self.root.deiconify()
    
    def load_image(self):
        file_path = filedialog.askopenfilename(
            title="Seleziona Immagine",
            filetypes=[("Immagini", "*.png *.jpg *.jpeg *.gif *.bmp"), ("Tutti i file", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    img_data = f.read()
                
                file_name = os.path.basename(file_path)
                
                self.current_attachments.append({
                    'name': file_name,
                    'type': 'image',
                    'data': img_data
                })
                
                self.update_attachments_preview()
                
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nel caricamento: {e}")
    
    def load_txt_file(self):
        file_path = filedialog.askopenfilename(
            title="Seleziona File TXT",
            filetypes=[("File di testo", "*.txt"), ("Tutti i file", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    file_data = f.read()
                
                file_name = os.path.basename(file_path)
                
                self.current_attachments.append({
                    'name': file_name,
                    'type': 'txt',
                    'data': file_data
                })
                
                self.update_attachments_preview()
                messagebox.showinfo("Successo", f"File TXT '{file_name}' caricato con successo!")
                
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nel caricamento del file TXT: {e}")
    
    def load_docx_file(self):
        file_path = filedialog.askopenfilename(
            title="Seleziona File DOCX",
            filetypes=[("File Word", "*.docx"), ("Tutti i file", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    file_data = f.read()
                
                file_name = os.path.basename(file_path)
                
                self.current_attachments.append({
                    'name': file_name,
                    'type': 'docx',
                    'data': file_data
                })
                
                self.update_attachments_preview()
                messagebox.showinfo("Successo", f"File DOCX '{file_name}' caricato con successo!")
                
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nel caricamento del file DOCX: {e}")
    
    def remove_attachment(self):
        if not self.current_attachments:
            messagebox.showwarning("Attenzione", "Nessun allegato da rimuovere!")
            return
        
        if len(self.current_attachments) == 1:
            self.current_attachments.pop(0)
        else:
            idx = messagebox.askquestion("Rimuovi", 
                f"Rimuovere l'ultimo allegato ({self.current_attachments[-1]['name']})?")
            if idx == 'yes':
                self.current_attachments.pop(-1)
        
        self.update_attachments_preview()
    
    def update_attachments_preview(self):
        for widget in self.preview_container.winfo_children():
            widget.destroy()
        
        self.preview_widgets = []
        
        for idx, attachment in enumerate(self.current_attachments):
            frame = ttk.LabelFrame(self.preview_container, text=f"{attachment['name']} [{attachment['type'].upper()}]", padding="5")
            frame.pack(fill='x', pady=5)
            
            if attachment['type'] == 'image':
                try:
                    image = Image.open(io.BytesIO(attachment['data']))
                    image.thumbnail((250, 250))
                    
                    photo = ImageTk.PhotoImage(image)
                    self.preview_widgets.append(photo)
                    
                    label = ttk.Label(frame, image=photo)
                    label.pack()
                    
                except Exception as e:
                    ttk.Label(frame, text=f"Errore visualizzazione: {e}").pack()
            
            elif attachment['type'] == 'txt':
                try:
                    content = attachment['data'].decode('utf-8', errors='ignore')
                    preview = content[:200] + "..." if len(content) > 200 else content
                    
                    text_widget = tk.Text(frame, height=6, wrap=tk.WORD, bg='#f0f0f0')
                    text_widget.insert('1.0', preview)
                    text_widget.config(state='disabled')
                    text_widget.pack(fill='x')
                    
                except Exception as e:
                    ttk.Label(frame, text=f"Errore lettura: {e}").pack()
            
            elif attachment['type'] == 'docx':
                info_frame = ttk.Frame(frame)
                info_frame.pack(fill='x', pady=5)
                
                size_kb = len(attachment['data']) / 1024
                ttk.Label(info_frame, text=f"📝 Documento Word", font=('Arial', 10, 'bold')).pack()
                ttk.Label(info_frame, text=f"Dimensione: {size_kb:.2f} KB").pack()
    
    def save_record(self):
        macchina = self.macchina_entry.get().strip()
        operatore = self.operatore_entry.get().strip()
        categoria = self.categoria_combo.get()
        problema = self.problema_text.get('1.0', tk.END).strip()
        soluzione = self.soluzione_text.get('1.0', tk.END).strip()
        
        if not macchina or not operatore or not problema or not soluzione:
            messagebox.showwarning("Campi Mancanti", "Compila tutti i campi obbligatori!")
            return
        
        data_ora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            self.cursor.execute('''
                INSERT INTO interventi (data_ora, macchina, operatore, categoria, problema, soluzione)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (data_ora, macchina, operatore, categoria, problema, soluzione))
            
            intervento_id = self.cursor.lastrowid
            
            for attachment in self.current_attachments:
                self.cursor.execute('''
                    INSERT INTO allegati (intervento_id, nome_file, tipo_file, contenuto)
                    VALUES (?, ?, ?, ?)
                ''', (intervento_id, attachment['name'], attachment['type'], attachment['data']))
            
            self.conn.commit()
            
            num_images = sum(1 for a in self.current_attachments if a['type'] == 'image')
            num_txt = sum(1 for a in self.current_attachments if a['type'] == 'txt')
            num_docx = sum(1 for a in self.current_attachments if a['type'] == 'docx')
            
            msg = f"Intervento salvato!\n\nAllegati:\n"
            if num_images > 0:
                msg += f"  🖼️ {num_images} immagine/i\n"
            if num_txt > 0:
                msg += f"  📄 {num_txt} file TXT\n"
            if num_docx > 0:
                msg += f"  📝 {num_docx} file DOCX\n"
            
            messagebox.showinfo("Successo", msg)
            
            self.clear_fields()
            self.load_all_records()
            
        except sqlite3.Error as e:
            messagebox.showerror("Errore Database", f"Errore: {e}")
    
    def clear_fields(self):
        self.macchina_entry.delete(0, tk.END)
        self.operatore_entry.delete(0, tk.END)
        self.categoria_combo.current(0)
        self.problema_text.delete('1.0', tk.END)
        self.soluzione_text.delete('1.0', tk.END)
        self.current_attachments = []
        self.update_attachments_preview()
    
    def load_all_records(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.cursor.execute('''
            SELECT id, data_ora, macchina, operatore, categoria, problema 
            FROM interventi 
            ORDER BY data_ora DESC
        ''')
        
        for row in self.cursor.fetchall():
            problema_short = row[5][:80] + '...' if len(row[5]) > 80 else row[5]
            self.tree.insert('', tk.END, iid=row[0], values=(row[1], row[2], row[3], row[4], problema_short))
    
    def search_records(self):
        search_term = self.search_entry.get().strip().lower()
        
        if not search_term:
            self.load_all_records()
            return
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.cursor.execute('''
            SELECT id, data_ora, macchina, operatore, categoria, problema 
            FROM interventi 
            WHERE LOWER(problema) LIKE ? OR LOWER(soluzione) LIKE ? OR LOWER(macchina) LIKE ?
            ORDER BY data_ora DESC
        ''', (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
        
        results = self.cursor.fetchall()
        
        if not results:
            messagebox.showinfo("Ricerca", "Nessun risultato trovato.")
            return
        
        for row in results:
            problema_short = row[5][:80] + '...' if len(row[5]) > 80 else row[5]
            self.tree.insert('', tk.END, iid=row[0], values=(row[1], row[2], row[3], row[4], problema_short))
    
    def show_details(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        
        record_id = selection[0]
        
        self.cursor.execute('''
            SELECT data_ora, macchina, operatore, categoria, problema, soluzione 
            FROM interventi 
            WHERE id = ?
        ''', (record_id,))
        
        record = self.cursor.fetchone()
        
        if record:
            self.cursor.execute('SELECT tipo_file, COUNT(*) FROM allegati WHERE intervento_id = ? GROUP BY tipo_file', (record_id,))
            attachments_count = dict(self.cursor.fetchall())
            
            num_images = attachments_count.get('image', 0)
            num_txt = attachments_count.get('txt', 0)
            num_docx = attachments_count.get('docx', 0)
            total_attachments = num_images + num_txt + num_docx
            
            attach_info = f"{total_attachments} ("
            parts = []
            if num_images > 0:
                parts.append(f"{num_images} img")
            if num_txt > 0:
                parts.append(f"{num_txt} txt")
            if num_docx > 0:
                parts.append(f"{num_docx} docx")
            attach_info += ", ".join(parts) + ")"
            
            details = f"""DATA/ORA: {record[0]}
MACCHINA: {record[1]}
OPERATORE: {record[2]}
CATEGORIA: {record[3]}
ALLEGATI: {attach_info}

PROBLEMA:
{record[4]}

SOLUZIONE:
{record[5]}"""
            
            self.details_text.config(state='normal')
            self.details_text.delete('1.0', tk.END)
            self.details_text.insert('1.0', details)
            self.details_text.config(state='disabled')
    
    def view_attachments(self):
        """Visualizza gli allegati dell'intervento selezionato"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Selezione", "Seleziona un intervento!")
            return
        
        record_id = selection[0]
        
        self.cursor.execute('SELECT nome_file, tipo_file, contenuto FROM allegati WHERE intervento_id = ?', (record_id,))
        attachments = self.cursor.fetchall()
        
        if not attachments:
            messagebox.showinfo("Allegati", "Nessun allegato per questo intervento.")
            return
        
        attach_window = tk.Toplevel(self.root)
        attach_window.title("Allegati Intervento")
        attach_window.geometry("900x700")
        
        notebook = ttk.Notebook(attach_window)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        images = [(nome, data) for nome, tipo, data in attachments if tipo == 'image']
        if images:
            img_tab = ttk.Frame(notebook)
            notebook.add(img_tab, text=f"🖼️ Immagini ({len(images)})")
            
            canvas = tk.Canvas(img_tab)
            scrollbar = ttk.Scrollbar(img_tab, orient="vertical", command=canvas.yview)
            container = ttk.Frame(canvas)
            
            container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=container, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            temp_photos = []
            
            for idx, (nome, img_data) in enumerate(images):
                frame = ttk.LabelFrame(container, text=nome, padding="10")
                frame.pack(fill='x', pady=10, padx=10)
                
                try:
                    image = Image.open(io.BytesIO(img_data))
                    image.thumbnail((800, 800))
                    
                    photo = ImageTk.PhotoImage(image)
                    temp_photos.append(photo)
                    
                    label = ttk.Label(frame, image=photo)
                    label.pack()
                    
                    btn_frame = ttk.Frame(frame)
                    btn_frame.pack(pady=5)
                    ttk.Button(btn_frame, text="💾 Salva Immagine", 
                             command=lambda d=img_data, n=nome: self.save_attachment_to_file(d, n)).pack()
                    
                except Exception as e:
                    ttk.Label(frame, text=f"Errore: {e}").pack()
            
            attach_window.photos = temp_photos
        
        txt_files = [(nome, data) for nome, tipo, data in attachments if tipo == 'txt']
        if txt_files:
            txt_tab = ttk.Frame(notebook)
            notebook.add(txt_tab, text=f"📄 File TXT ({len(txt_files)})")
            
            for idx, (nome, file_data) in enumerate(txt_files):
                frame = ttk.LabelFrame(txt_tab, text=nome, padding="10")
                frame.pack(fill='both', expand=True, padx=10, pady=5)
                
                try:
                    content = file_data.decode('utf-8', errors='ignore')
                    
                    text_widget = scrolledtext.ScrolledText(frame, wrap=tk.WORD, height=20)
                    text_widget.insert('1.0', content)
                    text_widget.config(state='disabled')
                    text_widget.pack(fill='both', expand=True)
                    
                    # Pulsanti
                    btn_frame = ttk.Frame(frame)
                    btn_frame.pack(pady=5)
                    ttk.Button(btn_frame, text="💾 Salva File", 
                             command=lambda d=file_data, n=nome: self.save_attachment_to_file(d, n)).pack(side=tk.LEFT, padx=5)
                    ttk.Button(btn_frame, text="📋 Copia Contenuto", 
                             command=lambda c=content: self.copy_to_clipboard(c)).pack(side=tk.LEFT, padx=5)
                    
                except Exception as e:
                    ttk.Label(frame, text=f"Errore lettura: {e}").pack()
        
        docx_files = [(nome, data) for nome, tipo, data in attachments if tipo == 'docx']
        if docx_files:
            docx_tab = ttk.Frame(notebook)
            notebook.add(docx_tab, text=f"📝 File DOCX ({len(docx_files)})")
            
            canvas = tk.Canvas(docx_tab)
            scrollbar = ttk.Scrollbar(docx_tab, orient="vertical", command=canvas.yview)
            container = ttk.Frame(canvas)
            
            container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=container, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            for idx, (nome, file_data) in enumerate(docx_files):
                frame = ttk.LabelFrame(container, text=nome, padding="15")
                frame.pack(fill='x', pady=10, padx=10)
                
                size_kb = len(file_data) / 1024
                
                info_frame = ttk.Frame(frame)
                info_frame.pack(fill='x', pady=10)
                
                ttk.Label(info_frame, text="📝", font=('Arial', 30)).pack()
                ttk.Label(info_frame, text=f"Documento Word", font=('Arial', 12, 'bold')).pack(pady=5)
                ttk.Label(info_frame, text=f"Dimensione: {size_kb:.2f} KB").pack()
                
                try:
                    import mammoth
                    result = mammoth.extract_raw_text(io.BytesIO(file_data))
                    text_content = result.value
                    
                    if text_content.strip():
                        ttk.Label(frame, text="Anteprima contenuto:", font=('Arial', 10, 'bold')).pack(pady=(10, 5))
                        
                        preview = text_content[:500] + "..." if len(text_content) > 500 else text_content
                        text_widget = scrolledtext.ScrolledText(frame, wrap=tk.WORD, height=10)
                        text_widget.insert('1.0', preview)
                        text_widget.config(state='disabled')
                        text_widget.pack(fill='x', pady=5)
                except:
                    ttk.Label(frame, text="(Installa 'mammoth' per vedere l'anteprima del contenuto)", 
                            font=('Arial', 9, 'italic')).pack(pady=5)
                
                btn_frame = ttk.Frame(frame)
                btn_frame.pack(pady=10)
                ttk.Button(btn_frame, text="💾 Salva File DOCX", 
                         command=lambda d=file_data, n=nome: self.save_attachment_to_file(d, n)).pack(side=tk.LEFT, padx=5)
                ttk.Button(btn_frame, text="📂 Apri con Word", 
                         command=lambda d=file_data, n=nome: self.open_docx_external(d, n)).pack(side=tk.LEFT, padx=5)
    
    def save_attachment_to_file(self, data, filename):
        file_path = filedialog.asksaveasfilename(
            defaultextension=os.path.splitext(filename)[1],
            initialfile=filename,
            filetypes=[("Tutti i file", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'wb') as f:
                    f.write(data)
                messagebox.showinfo("Successo", f"File salvato in:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nel salvataggio: {e}")
    
    def copy_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("Copiato", "Contenuto copiato negli appunti!")
    
    def open_docx_external(self, data, filename):
        import tempfile
        import subprocess
        import sys
        
        try:
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, filename)
            
            with open(temp_path, 'wb') as f:
                f.write(data)
            
            if sys.platform == 'win32':
                os.startfile(temp_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', temp_path])
            else:
                subprocess.run(['xdg-open', temp_path])
                
            messagebox.showinfo("Apertura", f"File aperto con l'applicazione predefinita.\n\nPercorso temporaneo:\n{temp_path}")
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile aprire il file: {e}")
    
    def delete_record(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Selezione", "Seleziona un intervento!")
            return
        
        if messagebox.askyesno("Conferma", "Eliminare questo intervento e tutti i suoi allegati?"):
            record_id = selection[0]
            
            try:
                self.cursor.execute('DELETE FROM allegati WHERE intervento_id = ?', (record_id,))
                self.cursor.execute('DELETE FROM interventi WHERE id = ?', (record_id,))
                self.conn.commit()
                
                messagebox.showinfo("Successo", "Intervento eliminato!")
                self.load_all_records()
                
                self.details_text.config(state='normal')
                self.details_text.delete('1.0', tk.END)
                self.details_text.config(state='disabled')
                
            except sqlite3.Error as e:
                messagebox.showerror("Errore", f"Errore: {e}")
    
    def ai_find_solutions(self):
        question = self.ai_question.get('1.0', tk.END).strip()
        
        if not question:
            messagebox.showwarning("Attenzione", "Inserisci una descrizione del problema!")
            return
        
        self.cursor.execute('SELECT id, problema, soluzione, macchina, categoria, data_ora FROM interventi')
        all_records = self.cursor.fetchall()
        
        if not all_records:
            messagebox.showinfo("IA", "Nessun intervento nel database.")
            return
        
        similarities = []
        for record in all_records:
            similarity = self.calculate_similarity(question.lower(), record[1].lower())
            if similarity > 0.3:  # Soglia minima
                similarities.append((similarity, record))
        
        similarities.sort(reverse=True, key=lambda x: x[0])
        
        self.ai_results.config(state='normal')
        self.ai_results.delete('1.0', tk.END)
        
        if not similarities:
            self.ai_results.insert('1.0', "❌ Nessuna soluzione simile trovata nel database.\n\n")
            self.ai_results.insert(tk.END, "Suggerimenti:\n")
            self.ai_results.insert(tk.END, "- Prova a descrivere il problema in modo diverso\n")
            self.ai_results.insert(tk.END, "- Usa parole chiave più generiche\n")
            self.ai_results.insert(tk.END, "- Aggiungi più interventi al database per migliorare i risultati")
        else:
            self.ai_results.insert('1.0', f"✅ Trovate {len(similarities[:5])} soluzioni simili:\n\n")
            self.ai_results.insert(tk.END, "="*80 + "\n\n")
            
            for idx, (similarity, record) in enumerate(similarities[:5]):
                record_id, problema, soluzione, macchina, categoria, data_ora = record
                percentage = int(similarity * 100)
                
                self.ai_results.insert(tk.END, f"🔍 RISULTATO #{idx+1} - Similarità: {percentage}%\n")
                self.ai_results.insert(tk.END, f"{'─'*80}\n")
                self.ai_results.insert(tk.END, f"📅 Data: {data_ora}\n")
                self.ai_results.insert(tk.END, f"🔧 Macchina: {macchina}\n")
                self.ai_results.insert(tk.END, f"📂 Categoria: {categoria}\n\n")
                self.ai_results.insert(tk.END, f"❓ PROBLEMA:\n{problema}\n\n")
                self.ai_results.insert(tk.END, f"✅ SOLUZIONE:\n{soluzione}\n\n")
                self.ai_results.insert(tk.END, "="*80 + "\n\n")
        
        self.ai_results.config(state='disabled')
    
    def calculate_similarity(self, text1, text2):
        return SequenceMatcher(None, text1, text2).ratio()
    
    def update_statistics(self):
        for widget in self.stats_container.winfo_children():
            widget.destroy()
        
        self.cursor.execute('SELECT COUNT(*) FROM interventi')
        total = self.cursor.fetchone()[0]
        
        if total == 0:
            ttk.Label(self.stats_container, text="Nessun dato disponibile", 
                     font=('Arial', 14)).pack(pady=50)
            return
        
        info_frame = ttk.LabelFrame(self.stats_container, text="Informazioni Generali", padding="15")
        info_frame.pack(fill='x', padx=10, pady=10)
        
        self.cursor.execute('SELECT COUNT(*) FROM allegati')
        total_attachments = self.cursor.fetchone()[0]
        
        self.cursor.execute('SELECT COUNT(*) FROM allegati WHERE tipo_file = "image"')
        total_images = self.cursor.fetchone()[0]
        
        self.cursor.execute('SELECT COUNT(*) FROM allegati WHERE tipo_file = "txt"')
        total_txt = self.cursor.fetchone()[0]
        
        self.cursor.execute('SELECT COUNT(*) FROM allegati WHERE tipo_file = "docx"')
        total_docx = self.cursor.fetchone()[0]
        
        self.cursor.execute('SELECT COUNT(DISTINCT macchina) FROM interventi')
        unique_machines = self.cursor.fetchone()[0]
        
        info_text = f"""
        📊 Totale Interventi: {total}
        📎 Totale Allegati: {total_attachments}
            🖼️ Immagini: {total_images}
            📄 File TXT: {total_txt}
            📝 File DOCX: {total_docx}
        🔧 Macchine Diverse: {unique_machines}
        📈 Media Allegati/Intervento: {total_attachments/total if total > 0 else 0:.1f}
        """
        
        ttk.Label(info_frame, text=info_text, font=('Arial', 11)).pack()
        
        charts_frame = ttk.Frame(self.stats_container)
        charts_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.cursor.execute('''
            SELECT categoria, COUNT(*) 
            FROM interventi 
            GROUP BY categoria 
            ORDER BY COUNT(*) DESC
        ''')
        cat_data = self.cursor.fetchall()
        
        if cat_data:
            fig1, ax1 = plt.subplots(figsize=(6, 4))
            categories = [row[0] for row in cat_data]
            counts = [row[1] for row in cat_data]
            
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#C7CEEA']
            ax1.bar(categories, counts, color=colors[:len(categories)])
            ax1.set_title('Interventi per Categoria', fontsize=14, fontweight='bold')
            ax1.set_xlabel('Categoria')
            ax1.set_ylabel('Numero Interventi')
            ax1.tick_params(axis='x', rotation=45)
            plt.tight_layout()
            
            canvas1 = FigureCanvasTkAgg(fig1, charts_frame)
            canvas1.draw()
            canvas1.get_tk_widget().pack(side=tk.LEFT, fill='both', expand=True, padx=5)
        
        self.cursor.execute('''
            SELECT macchina, COUNT(*) as cnt
            FROM interventi 
            GROUP BY macchina 
            ORDER BY cnt DESC 
            LIMIT 5
        ''')
        machine_data = self.cursor.fetchall()
        
        if machine_data:
            fig2, ax2 = plt.subplots(figsize=(6, 4))
            machines = [row[0][:15] + '...' if len(row[0]) > 15 else row[0] for row in machine_data]
            counts = [row[1] for row in machine_data]
            
            ax2.barh(machines, counts, color='#FF6B6B')
            ax2.set_title('Top 5 Macchine - Interventi', fontsize=14, fontweight='bold')
            ax2.set_xlabel('Numero Interventi')
            ax2.invert_yaxis()
            plt.tight_layout()
            
            canvas2 = FigureCanvasTkAgg(fig2, charts_frame)
            canvas2.draw()
            canvas2.get_tk_widget().pack(side=tk.LEFT, fill='both', expand=True, padx=5)
        
        self.cursor.execute('''
            SELECT strftime('%Y-%m', data_ora) as mese, COUNT(*) 
            FROM interventi 
            GROUP BY mese 
            ORDER BY mese DESC 
            LIMIT 12
        ''')
        month_data = self.cursor.fetchall()
        
        if month_data and len(month_data) > 1:
            fig3, ax3 = plt.subplots(figsize=(12, 4))
            months = [row[0] for row in reversed(month_data)]
            counts = [row[1] for row in reversed(month_data)]
            
            ax3.plot(months, counts, marker='o', linewidth=2, markersize=8, color='#4ECDC4')
            ax3.fill_between(range(len(months)), counts, alpha=0.3, color='#4ECDC4')
            ax3.set_title('Trend Interventi per Mese', fontsize=14, fontweight='bold')
            ax3.set_xlabel('Mese')
            ax3.set_ylabel('Numero Interventi')
            ax3.tick_params(axis='x', rotation=45)
            ax3.grid(True, alpha=0.3)
            plt.tight_layout()
            
            canvas3 = FigureCanvasTkAgg(fig3, self.stats_container)
            canvas3.draw()
            canvas3.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
    
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"export_interventi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            self.cursor.execute('''
                SELECT id, data_ora, macchina, operatore, categoria, problema, soluzione 
                FROM interventi 
                ORDER BY data_ora DESC
            ''')
            records = self.cursor.fetchall()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Interventi"
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            headers = ['ID', 'Data/Ora', 'Macchina', 'Operatore', 'Categoria', 'Problema', 'Soluzione', 'Allegati']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row_idx, record in enumerate(records, 2):
                self.cursor.execute('''
                    SELECT tipo_file, COUNT(*) 
                    FROM allegati 
                    WHERE intervento_id = ? 
                    GROUP BY tipo_file
                ''', (record[0],))
                attach_counts = dict(self.cursor.fetchall())
                
                attach_str = []
                if attach_counts.get('image', 0) > 0:
                    attach_str.append(f"{attach_counts['image']} img")
                if attach_counts.get('txt', 0) > 0:
                    attach_str.append(f"{attach_counts['txt']} txt")
                if attach_counts.get('docx', 0) > 0:
                    attach_str.append(f"{attach_counts['docx']} docx")
                
                for col_idx, value in enumerate(record, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                ws.cell(row=row_idx, column=8).value = ", ".join(attach_str) if attach_str else "Nessuno"
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 50
            ws.column_dimensions['H'].width = 20
            
            ws_stats = wb.create_sheet("Statistiche")
            
            ws_stats['A1'] = 'STATISTICHE PER CATEGORIA'
            ws_stats['A1'].font = Font(bold=True, size=14)
            ws_stats['A3'] = 'Categoria'
            ws_stats['B3'] = 'Conteggio'
            
            self.cursor.execute('SELECT categoria, COUNT(*) FROM interventi GROUP BY categoria ORDER BY COUNT(*) DESC')
            for idx, (cat, count) in enumerate(self.cursor.fetchall(), 4):
                ws_stats[f'A{idx}'] = cat
                ws_stats[f'B{idx}'] = count
            
            start_row = idx + 3
            ws_stats[f'A{start_row}'] = 'TOP 10 MACCHINE'
            ws_stats[f'A{start_row}'].font = Font(bold=True, size=14)
            ws_stats[f'A{start_row+2}'] = 'Macchina'
            ws_stats[f'B{start_row+2}'] = 'Interventi'
            
            self.cursor.execute('SELECT macchina, COUNT(*) FROM interventi GROUP BY macchina ORDER BY COUNT(*) DESC LIMIT 10')
            for idx, (machine, count) in enumerate(self.cursor.fetchall(), start_row+3):
                ws_stats[f'A{idx}'] = machine
                ws_stats[f'B{idx}'] = count
            
            ws_stats.column_dimensions['A'].width = 30
            ws_stats.column_dimensions['B'].width = 15
            
            wb.save(file_path)
            messagebox.showinfo("Successo", f"Dati esportati con successo!\n\n{len(records)} interventi salvati in:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Errore Export", f"Errore durante l'export: {e}")
    
    def __del__(self):
        """Chiude connessione database"""
        if hasattr(self, 'conn'):
            self.conn.close()

def main():
    root = tk.Tk()
    app = MachineTrackerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()